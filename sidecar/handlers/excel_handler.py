"""Excel 文档处理器
基于 openpyxl 实现 Excel 文档的读取、转换、分析
精简版：仅支持 read/convert/analyze 操作
"""

import os
import csv
import io
import html
import logging

from openpyxl import load_workbook


class ExcelHandler:
    """Excel (.xlsx) 文档处理器（精简版，仅支持 read/convert/analyze）"""

    logger = logging.getLogger(__name__)

    def read(self, params: dict) -> dict:
        """读取 Excel 文档

        params:
            path: 文件路径
            sheet: 工作表名称（可选，默认读取所有）
            range: 读取范围（可选，如 "A1:D10"）
            include_formatting: 是否提取单元格格式（字体/填充/边框/对齐/数字格式）
            include_formulas: 是否分离公式与计算结果值（同时加载 data_only=True 工作簿）
            include_charts: 是否提取图表信息
            include_merged_cells: 是否提取合并单元格范围列表
            include_comments: 是否提取单元格批注
        """
        path = params.get("path", "")
        sheet_name = params.get("sheet", None)
        read_range = params.get("range", None)
        if not path:
            self.logger.error("read: 缺少文件路径")
            return {"error": "缺少文件路径"}
        if not os.path.exists(path):
            raise FileNotFoundError(path)

        self.logger.info("read: 开始读取 Excel 文档, path=%s", path)

        # 解析扩展参数
        include_formatting = params.get("include_formatting", False)
        include_formulas = params.get("include_formulas", False)
        include_charts = params.get("include_charts", False)
        include_merged_cells = params.get("include_merged_cells", False)
        include_comments = params.get("include_comments", False)

        # 公式分离需要同时加载 data_only=True 工作簿以获取计算结果值
        wb = load_workbook(path, data_only=False)
        wb_values = None
        if include_formulas:
            try:
                wb_values = load_workbook(path, data_only=True)
            except Exception as e:
                self.logger.warning("read: 加载 data_only 工作簿失败，公式值分离将不可用: %s", e)
                wb_values = None

        result = {"sheets": {}}

        if sheet_name:
            sheets_to_read = [sheet_name]
        else:
            sheets_to_read = wb.sheetnames

        for name in sheets_to_read:
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            # 对应的 data_only 工作表（用于公式值分离）
            ws_values = wb_values[name] if wb_values and name in wb_values.sheetnames else None

            # ------------------------------------------------------------------ #
            #  读取单元格数据
            # ------------------------------------------------------------------ #
            if read_range:
                rows = []
                for row in ws[read_range]:
                    row_data = []
                    for cell in row:
                        # 获取 data_only 工作簿中对应位置的单元格
                        value_cell = None
                        if ws_values:
                            try:
                                value_cell = ws_values.cell(row=cell.row, column=cell.column)
                            except Exception:
                                value_cell = None
                        row_data.append(self._extract_cell_value(
                            cell, value_cell, include_formulas, include_formatting, include_comments
                        ))
                    rows.append(row_data)
            else:
                rows = []
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
                    row_data = []
                    for cell in row:
                        # 获取 data_only 工作簿中对应位置的单元格
                        value_cell = None
                        if ws_values:
                            try:
                                value_cell = ws_values.cell(row=cell.row, column=cell.column)
                            except Exception:
                                value_cell = None
                        row_data.append(self._extract_cell_value(
                            cell, value_cell, include_formulas, include_formatting, include_comments
                        ))
                    rows.append(row_data)

            sheet_info = {
                "data": rows,
                "row_count": ws.max_row,
                "col_count": ws.max_column,
            }

            # ------------------------------------------------------------------ #
            #  合并单元格范围列表
            # ------------------------------------------------------------------ #
            if include_merged_cells:
                merged_ranges = []
                try:
                    for merged_range in ws.merged_cells.ranges:
                        merged_ranges.append(str(merged_range))
                except Exception as e:
                    self.logger.warning("read: 提取合并单元格失败: %s", e)
                sheet_info["merged_cells"] = merged_ranges

            # ------------------------------------------------------------------ #
            #  图表信息
            # ------------------------------------------------------------------ #
            if include_charts:
                charts_info = []
                try:
                    for chart in ws._charts:
                        chart_info = {
                            # 图表类型（BarChart/LineChart/PieChart 等）
                            "type": type(chart).__name__,
                            # 标题（可能为 None）
                            "title": self._extract_chart_title(chart),
                        }
                        charts_info.append(chart_info)
                except Exception as e:
                    self.logger.warning("read: 提取图表信息失败: %s", e)
                sheet_info["charts"] = charts_info

            result["sheets"][name] = sheet_info

        result["sheet_names"] = wb.sheetnames
        self.logger.info("read: Excel 文档读取完成, path=%s, 工作表数=%d", path, len(result["sheets"]))
        return result

    @staticmethod
    def _extract_cell_value(cell, value_cell, include_formulas: bool, include_formatting: bool, include_comments: bool) -> dict:
        """提取单元格值，可选包含公式、格式、批注

        Args:
            cell: 原始单元格（data_only=False，显示公式）
            value_cell: data_only=True 工作簿中对应位置的单元格（显示计算结果值）
            include_formulas: 是否分离公式与值
            include_formatting: 是否提取单元格格式
            include_comments: 是否提取批注
        """
        # 简化模式：仅返回值（向后兼容）
        if not include_formulas and not include_formatting and not include_comments:
            val = cell.value
            if val is None and cell.data_type == "f":
                val = cell.internal_value
            return val

        # 详细模式：返回字典结构
        cell_info = {}

        # 值与公式分离
        if include_formulas:
            # data_type == 'f' 表示该单元格包含公式
            if cell.data_type == "f":
                cell_info["formula"] = cell.value
                # 计算结果值来自 data_only 工作簿
                cell_info["value"] = value_cell.value if value_cell else None
            else:
                cell_info["value"] = cell.value
                cell_info["formula"] = None
        else:
            val = cell.value
            if val is None and cell.data_type == "f":
                val = cell.internal_value
            cell_info["value"] = val

        # 单元格格式
        if include_formatting:
            cell_info["formatting"] = ExcelHandler._extract_cell_formatting(cell)

        # 批注
        if include_comments:
            if cell.comment:
                cell_info["comment"] = {
                    "text": cell.comment.text,
                    "author": cell.comment.author or "",
                }
            else:
                cell_info["comment"] = None

        return cell_info

    @staticmethod
    def _extract_cell_formatting(cell) -> dict:
        """提取单元格格式信息（字体/填充/边框/对齐/数字格式）"""
        fmt = {}

        # 字体
        font = cell.font
        if font:
            fmt["font"] = {
                "name": font.name,
                "size": font.size,
                "bold": font.bold,
                "italic": font.italic,
                "underline": font.underline,
                # 颜色：可能是 RGBColor 或 ThemeColor
                "color": str(font.color.rgb) if font.color and font.color.rgb else None,
            }

        # 填充
        fill = cell.fill
        if fill:
            # openpyxl 不同版本的 PatternFill 属性名不一致：
            # 新版为 fill_type，旧版为 patternType；pattern_type（snake_case）是无效属性
            fill_info = {"pattern_type": getattr(fill, "fill_type", None) or getattr(fill, "patternType", None)}
            if fill.fgColor and fill.fgColor.rgb:
                fill_info["fg_color"] = str(fill.fgColor.rgb)
            if fill.bgColor and fill.bgColor.rgb:
                fill_info["bg_color"] = str(fill.bgColor.rgb)
            fmt["fill"] = fill_info

        # 边框
        border = cell.border
        if border:
            fmt["border"] = {
                "left_style": border.left.style if border.left else None,
                "right_style": border.right.style if border.right else None,
                "top_style": border.top.style if border.top else None,
                "bottom_style": border.bottom.style if border.bottom else None,
            }

        # 对齐
        alignment = cell.alignment
        if alignment:
            fmt["alignment"] = {
                "horizontal": alignment.horizontal,
                "vertical": alignment.vertical,
                "wrap_text": alignment.wrap_text,
                "indent": alignment.indent,
            }

        # 数字格式
        if cell.number_format:
            fmt["number_format"] = cell.number_format

        return fmt

    @staticmethod
    def _extract_chart_title(chart) -> str:
        """尝试从图表对象提取标题，失败返回空字符串"""
        try:
            if hasattr(chart, "title") and chart.title:
                # 标题可能是 RichText 或字符串
                if hasattr(chart.title, "tx") and chart.title.tx:
                    if hasattr(chart.title.tx, "rich") and chart.title.tx.rich:
                        paragraphs = chart.title.tx.rich.paragraphs
                        if paragraphs:
                            runs = paragraphs[0].runs
                            if runs:
                                return runs[0].text
                return str(chart.title)
        except Exception:
            pass
        return ""

    def convert(self, params: dict) -> dict:
        """格式转换

        params:
            path: 源文件路径
            output_path: 输出文件路径（可选）
            format: 目标格式（csv, pdf, html, txt）
            sheet: 工作表名称（可选，默认转换所有）
        """
        path = params.get("path", "")
        output_path = params.get("output_path", "")
        target_format = params.get("format", "").lower()
        sheet_name = params.get("sheet", None)

        if not path:
            self.logger.error("convert: 缺少源文件路径")
            return {"error": "缺少源文件路径"}
        if not os.path.exists(path):
            raise FileNotFoundError(path)
        if target_format not in ("csv", "pdf", "html", "txt"):
            self.logger.error("convert: 不支持的目标格式: %s", target_format)
            return {"error": f"不支持的目标格式: {target_format}，支持的格式: csv, pdf, html, txt"}

        self.logger.info("convert: 开始格式转换, path=%s, format=%s", path, target_format)

        wb = load_workbook(path, data_only=True)

        # 确定要转换的工作表
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                self.logger.error("convert: 工作表不存在: %s", sheet_name)
                return {"error": f"工作表不存在: {sheet_name}"}
            sheets_to_convert = [sheet_name]
        else:
            sheets_to_convert = wb.sheetnames

        # 读取所有工作表数据
        all_sheets_data = {}
        for name in sheets_to_convert:
            ws = wb[name]
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
                rows.append([cell if cell is not None else "" for cell in row])
            all_sheets_data[name] = rows

        # PDF 为二进制格式，必须写入文件
        if target_format == "pdf":
            if not output_path:
                base, _ = os.path.splitext(path)
                output_path = base + ".pdf"
            self._convert_to_pdf(all_sheets_data, output_path)
            self.logger.info("convert: 格式转换完成, output_path=%s, format=%s", output_path, target_format)
            return {
                "path": output_path,
                "format": target_format,
                "message": f"已转换为 {target_format} 格式",
            }

        # 根据目标格式生成文本内容
        if target_format == "csv":
            content = self._convert_to_csv(all_sheets_data)
        elif target_format == "html":
            content = self._convert_to_html(all_sheets_data)
        elif target_format == "txt":
            content = self._convert_to_txt(all_sheets_data)

        # 写入输出文件或返回内容
        if output_path:
            os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(content)
            self.logger.info("convert: 格式转换完成, output_path=%s, format=%s", output_path, target_format)
            return {
                "path": output_path,
                "format": target_format,
                "message": f"已转换为 {target_format} 格式",
            }
        else:
            return {
                "content": content,
                "format": target_format,
            }

    def analyze(self, params: dict) -> dict:
        """分析 Excel 文档"""
        path = params.get("path", "")
        if not path:
            self.logger.error("analyze: 缺少文件路径")
            return {"error": "缺少文件路径"}
        if not os.path.exists(path):
            raise FileNotFoundError(path)

        self.logger.info("analyze: 开始分析 Excel 文档, path=%s", path)

        wb = load_workbook(path, data_only=True)
        sheets_info = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheets_info.append({
                "name": name,
                "rows": ws.max_row,
                "cols": ws.max_column,
            })

        self.logger.info("analyze: Excel 文档分析完成, path=%s, 工作表数=%d", path, len(wb.sheetnames))
        return {
            "file_size": os.path.getsize(path),
            "sheet_count": len(wb.sheetnames),
            "sheets": sheets_info,
        }

    # ------------------------------------------------------------------ #
    #  格式转换辅助方法
    # ------------------------------------------------------------------ #

    def _convert_to_csv(self, all_sheets_data: dict) -> str:
        """将工作表数据转换为 CSV 格式"""
        parts = []
        for sheet_name, rows in all_sheets_data.items():
            if len(all_sheets_data) > 1:
                parts.append(f"# 工作表: {sheet_name}")
            output = io.StringIO()
            writer = csv.writer(output)
            for row in rows:
                writer.writerow(row)
            parts.append(output.getvalue().rstrip("\r\n"))

        return "\n".join(parts)

    def _convert_to_pdf(self, all_sheets_data: dict, output_path: str):
        """将工作表数据转换为 PDF 格式（使用 reportlab 渲染表格）"""
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        except ImportError:
            self.logger.error("convert: reportlab 未安装，无法转换为 PDF")
            raise RuntimeError("reportlab 未安装，无法转换为 PDF")

        from handlers.font_utils import register_chinese_font
        font_name = register_chinese_font()

        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

        doc = SimpleDocTemplate(output_path, pagesize=A4)
        styles = getSampleStyleSheet()

        sheet_title_style = ParagraphStyle(
            "SheetTitle",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=14,
            spaceAfter=10,
            spaceBefore=20,
        )

        elements = []
        for sheet_name, rows in all_sheets_data.items():
            elements.append(Paragraph(html.escape(sheet_name), sheet_title_style))
            elements.append(Spacer(1, 0.5 * cm))

            if rows:
                table_data = [[str(cell) for cell in row] for row in rows]

                table = Table(table_data)
                table.setStyle(TableStyle([
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTNAME", (0, 0), (-1, -1), font_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
                    ("FONTNAME", (0, 0), (-1, 0), font_name),
                    ("FONTSIZE", (0, 0), (-1, 0), 9),
                ]))
                elements.append(table)
                elements.append(Spacer(1, 1 * cm))

        doc.build(elements)

    def _convert_to_html(self, all_sheets_data: dict) -> str:
        """将工作表数据转换为 HTML 表格"""
        parts = [
            "<!DOCTYPE html>",
            "<html>",
            "<head>",
            '<meta charset="utf-8">',
            "<title>Excel 转换</title>",
            "<style>",
            "body { font-family: sans-serif; margin: 20px; }",
            "table { border-collapse: collapse; margin-bottom: 20px; width: 100%; }",
            "th, td { border: 1px solid #ccc; padding: 6px 10px; text-align: left; }",
            "th { background-color: #f0f0f0; font-weight: bold; }",
            "h2 { margin-top: 30px; color: #333; }",
            "</style>",
            "</head>",
            "<body>",
        ]

        for sheet_name, rows in all_sheets_data.items():
            parts.append(f"<h2>{html.escape(sheet_name)}</h2>")
            parts.append("<table>")
            for i, row in enumerate(rows):
                tag = "th" if i == 0 else "td"
                parts.append("<tr>")
                for cell in row:
                    parts.append(f"<{tag}>{html.escape(str(cell))}</{tag}>")
                parts.append("</tr>")
            parts.append("</table>")

        parts.extend(["</body>", "</html>"])
        return "\n".join(parts)

    def _convert_to_txt(self, all_sheets_data: dict) -> str:
        """将工作表数据转换为纯文本（制表符分隔）"""
        parts = []
        for sheet_name, rows in all_sheets_data.items():
            if len(all_sheets_data) > 1:
                parts.append(f"=== {sheet_name} ===")
            for row in rows:
                parts.append("\t".join(str(cell) for cell in row))
            parts.append("")

        return "\n".join(parts)
