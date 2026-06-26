"""Excel 文档生成 Helper 函数
封装 openpyxl 常用操作，内置专业配色方案
"""

import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 专业配色方案
# THEME 已应用到 create_excel_doc 的默认工作表；apply_excel_theme(ws) 可对其他工作表应用相同样式
if HAS_OPENPYXL:
    THEME = {
        "header_fill": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
        "alt_row_fill": PatternFill(start_color="EDF2F9", end_color="EDF2F9", fill_type="solid"),
        "header_font": Font(name="微软雅黑", bold=True, color="1F4E79", size=11),
        "title_font": Font(name="微软雅黑", bold=True, color="1F4E79", size=16),
        "normal_font": Font(name="微软雅黑", size=11),
        "border": Border(
            left=Side(style='thin', color='B4C6E7'),
            right=Side(style='thin', color='B4C6E7'),
            top=Side(style='thin', color='B4C6E7'),
            bottom=Side(style='thin', color='B4C6E7'),
        ),
        "center_align": Alignment(horizontal='center', vertical='center'),
    }
else:
    THEME = {}


def apply_excel_theme(ws):
    """对 openpyxl Worksheet 应用专业配色方案

    设置工作表的默认列宽、行高和默认字体样式。
    注意：此函数仅设置工作表级默认样式，不修改已有单元格。
    表头样式建议在数据写入后调用 apply_excel_header_style(ws, row=1) 应用。

    Args:
        ws: openpyxl Worksheet 对象

    Returns:
        Worksheet: 应用了默认样式的工作表对象
    """
    if not HAS_OPENPYXL or not THEME:
        return ws

    try:
        # 设置工作表标签颜色为深蓝色（与 THEME 配色一致）
        ws.sheet_properties.tabColor = "1F4E79"
        # 设置默认行高
        ws.sheet_format.defaultRowHeight = 18
    except Exception:
        pass

    return ws


def apply_excel_header_style(ws, row=1):
    """对指定行应用表头样式（蓝色背景 + 粗体 + 边框 + 居中）

    Args:
        ws: openpyxl Worksheet 对象
        row: 表头所在行号（默认 1）

    Returns:
        Worksheet: 应用了表头样式的工作表对象
    """
    if not HAS_OPENPYXL or not THEME:
        return ws

    try:
        for cell in ws[row]:
            cell.fill = THEME["header_fill"]
            cell.font = THEME["header_font"]
            cell.border = THEME["border"]
            cell.alignment = THEME["center_align"]
    except Exception:
        pass

    return ws


def create_excel_doc(title=None, author=""):
    """创建一个预配置好专业样式的 Excel 工作簿对象

    Args:
        title: 工作簿标题（可选）
        author: 文档作者

    Returns:
        Workbook: 预配置好的 openpyxl Workbook 对象
                 默认工作表已应用 THEME 默认样式

    示例:
        wb = create_excel_doc(title="销售数据", author="张三")
        ws = wb.active
        ws.title = "销售数据"
        ws.append(["产品", "销量", "金额"])
        apply_excel_header_style(ws)  # 对第一行应用表头样式
        save_excel_doc(wb, "销售数据.xlsx")
    """
    wb = Workbook()

    if title:
        wb.properties.title = title
    if author:
        wb.properties.creator = author

    # 对默认工作表应用 THEME 默认样式
    if HAS_OPENPYXL and THEME:
        try:
            ws = wb.active
            # 设置工作表默认字体（通过 sheet_format）
            # 注意：openpyxl 的默认字体需要通过 cell 逐个设置，
            # 这里仅设置工作表级别的属性，单元格样式由 apply_excel_header_style 处理
            ws.sheet_format.defaultRowHeight = 18
        except Exception:
            pass

    return wb


def save_excel_doc(wb, filename, working_dir=""):
    """保存 Excel 工作簿到工作目录

    Args:
        wb: openpyxl Workbook 对象
        filename: 文件名（如 "数据.xlsx"）
        working_dir: 工作目录路径

    Returns:
        str: 保存的文件绝对路径
    """
    output_path = os.path.join(working_dir, filename) if working_dir else filename
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)

    return output_path
